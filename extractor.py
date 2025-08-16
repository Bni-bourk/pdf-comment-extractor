import fitz  # PyMuPDF
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import traceback
import openpyxl
import shutil
import sys
import re
from openpyxl.styles import Border, Side, Font, Alignment

# Determine base path (for PyInstaller or script)
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(".")

# Set template file name
template_file_name = "CRS.xlsx"
template_path = os.path.join(base_path, template_file_name)

# Check that the template file exists
if not os.path.exists(template_path):
    messagebox.showerror("❌ Error", f"Template file '{template_file_name}' not found in application folder.")
    sys.exit(1)

# Split text into sentences with line breaks
def split_sentences_with_linebreaks(text):
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z0-9])', text.strip())
    return "\n".join(s.strip() for s in sentences if s.strip())

# Extract field values from PDF text
def find_field(label, text):
    match = re.search(rf"{re.escape(label)}\s*[:：]?\s*(.+)", text, re.IGNORECASE)
    return match.group(1).strip() if match else ""

# Main extraction function
def extract_comments_to_crs_template(pdf_path, output_path):
    try:
        # Copy template to destination
        shutil.copy(template_path, output_path)
    except Exception as e:
        raise RuntimeError(f"Error copying template: {e}")

    comments = []
    try:
        with fitz.open(pdf_path) as doc:
            full_text = ""
            for i, page in enumerate(doc):
                full_text += page.get_text()

                annot = page.first_annot
                while annot:
                    if annot.type[1] == 'FreeText':
                        info = annot.info
                        comments.append([
                            i + 1,  # Page number
                            info.get("content", "N/A")
                        ])
                    annot = annot.next

        # Extract fields
        client_name = find_field("CLIENT NAME", full_text)
        project_description = find_field("PROJECT DESCRIPTION", full_text)
        project_number = find_field("PROJECT NUMBER", full_text)
        po_reference = find_field("PURCHASE ORDER REFERENCE", full_text)

    except Exception as e:
        raise RuntimeError("Error reading PDF: " + str(e))

    try:
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Fill extracted fields
        ws["B6"] = client_name
        ws["B7"] = project_description
        ws["B8"] = project_number
        ws["B9"] = po_reference

        # PDF name into B10
        pdf_filename = os.path.splitext(os.path.basename(pdf_path))[0]
        ws["B10"] = pdf_filename
        ws["B10"].font = Font(bold=True)

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 100

        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Insert comments
        start_row = 12
        start_col = 1  # A

        for idx, (page_num, comment) in enumerate(comments):
            row = start_row + idx
            formatted_comment = split_sentences_with_linebreaks(comment)

            # Insert data in A and B
            cell1 = ws.cell(row=row, column=start_col, value=page_num)
            cell2 = ws.cell(row=row, column=start_col + 1, value=formatted_comment)

            cell1.alignment = Alignment(wrap_text=True)
            cell2.alignment = Alignment(wrap_text=True)

            # Optional: clear other columns C to J if needed
            for col in range(3, 11):  # Columns C(3) to J(10)
                ws.cell(row=row, column=col, value="")

            # Apply borders across columns A to J for this row
            for col in range(1, 11):  # Columns A(1) to J(10)
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        wb.save(output_path)

        if comments:
            messagebox.showinfo("✅ Done", f"Extraction complete.\nSaved to:\n{output_path}")
            if messagebox.askyesno("Open File", "Do you want to open the Excel file now?"):
                os.startfile(output_path)
        else:
            messagebox.showinfo("ℹ️ No Comments Found", "No FreeText comments found in this PDF.")

    except Exception as e:
        raise RuntimeError("Error saving Excel file: " + str(e))

# GUI handler
def run_extraction():
    pdf_path = filedialog.askopenfilename(title="Select PDF File", filetypes=[("PDF Files", "*.pdf")])
    if not pdf_path:
        return

    pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
    default_excel_name = pdf_name + "_comments.xlsx"

    output_path = filedialog.asksaveasfilename(
        initialfile=default_excel_name,
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
        title="Save Excel File As"
    )
    if not output_path:
        return

    try:
        status_label.config(text="Processing...")
        root.update()
        extract_comments_to_crs_template(pdf_path, output_path)
        status_label.config(text="Done.")
    except Exception as e:
        messagebox.showerror("❌ Error", f"{str(e)}\n\nDetails:\n{traceback.format_exc()}")
        status_label.config(text="Failed.")

# GUI setup
root = tk.Tk()
root.title("PDF Comment Extractor")
root.geometry("420x250")
root.resizable(False, False)

label = tk.Label(root, text="📄 PDF Comment Extractor", font=("Arial", 13, "bold"))
label.pack(pady=(20, 5))

author_label = tk.Label(root, text="By Ayoub BNI-BOURK", font=("Arial", 9), fg="gray")
author_label.pack(pady=(0, 15))

extract_button = tk.Button(root, text="Select PDF and Extract", font=("Arial", 11), command=run_extraction)
extract_button.pack(pady=10, ipadx=10, ipady=5)

status_label = tk.Label(root, text="", font=("Arial", 9), fg="blue")
status_label.pack(pady=(5, 10))

root.mainloop()
